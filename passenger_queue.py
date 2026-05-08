"""
passenger_queue.py — Data-driven passenger queue and eVTOL boarding logic.

Reads passenger arrival times from the Tampa Bay AAM study Excel file.
Implements boarding rules:
  1. Passenger reneges if waiting > 30 min
  2. If first boarded pax waits 15 min with no second pax -> takeoff
  3. If 2+ pax boarded and second pax waited > 10 min -> takeoff
  4. eVTOL full at 4 pax -> immediate takeoff
"""

import os

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Boarding rule thresholds (minutes)
RENEGE_TIMEOUT_M = 30.0       # passenger leaves after 30 min wait
FIRST_PAX_TIMEOUT_M = 15.0    # first pax alone -> takeoff after 15 min
SECOND_PAX_TIMEOUT_M = 10.0   # 2+ pax -> takeoff 10 min after 2nd boards
EVTOL_CAPACITY = 4

# Data file — look in project data/ dir first, fall back to ~/Downloads
_HERE = os.path.dirname(os.path.abspath(__file__))
_EXCEL_PATH = os.path.join(_HERE, "data", "evtol_in_and_out_of_vehicle.xlsx")
if not os.path.isfile(_EXCEL_PATH):
    _EXCEL_PATH = os.path.join(
        os.path.expanduser("~"), "Downloads", "evtol in and out of vehicle.xlsx"
    )


class Passenger:
    __slots__ = ("pax_id", "origin_vp", "dest_vp", "arrival_time_s",
                 "board_time_s", "takeoff_time_s", "reneged")

    def __init__(self, pax_id: str, origin_vp: str, dest_vp: str,
                 arrival_time_s: float) -> None:
        self.pax_id = pax_id
        self.origin_vp = origin_vp
        self.dest_vp = dest_vp
        self.arrival_time_s = arrival_time_s
        self.board_time_s: float | None = None
        self.takeoff_time_s: float | None = None
        self.reneged = False


class VertiportQueue:
    """Passenger queue for one vertiport, one destination."""

    def __init__(self, origin_vp: str, dest_vp: str,
                 arrival_times_min: list[float]) -> None:
        self.origin_vp = origin_vp
        self.dest_vp = dest_vp

        # Convert arrival times to seconds and sort
        self._schedule_s = sorted(t * 60.0 for t in arrival_times_min)
        self._schedule_idx = 0
        self._pax_seq = 0

        # Waiting queue (not yet boarded)
        self.waiting: list[Passenger] = []
        # Currently boarding the eVTOL
        self.boarded: list[Passenger] = []

        # Completed records
        self.departed: list[Passenger] = []
        self.reneged_list: list[Passenger] = []

    @property
    def total_arrived(self) -> int:
        return self._pax_seq

    def step(self, t: float) -> dict | None:
        """
        Advance one simulation step. Returns a dispatch dict if eVTOL should
        take off, or None.
        """
        # 1. Inject newly arrived passengers
        while (self._schedule_idx < len(self._schedule_s)
               and self._schedule_s[self._schedule_idx] <= t):
            self._pax_seq += 1
            pax = Passenger(
                pax_id=f"pax_{self.origin_vp}_{self.dest_vp}_{self._pax_seq:04d}",
                origin_vp=self.origin_vp,
                dest_vp=self.dest_vp,
                arrival_time_s=self._schedule_s[self._schedule_idx],
            )
            self.waiting.append(pax)
            self._schedule_idx += 1

        # 2. Remove reneged passengers (waited > 30 min in queue)
        still_waiting = []
        for pax in self.waiting:
            wait_min = (t - pax.arrival_time_s) / 60.0
            if wait_min >= RENEGE_TIMEOUT_M:
                pax.reneged = True
                self.reneged_list.append(pax)
            else:
                still_waiting.append(pax)
        self.waiting = still_waiting

        # 3. Board waiting passengers into eVTOL (up to capacity)
        while self.waiting and len(self.boarded) < EVTOL_CAPACITY:
            pax = self.waiting.pop(0)
            pax.board_time_s = t
            self.boarded.append(pax)

        # 4. Check takeoff conditions
        if not self.boarded:
            return None

        n_boarded = len(self.boarded)
        first_board_t = self.boarded[0].board_time_s
        time_since_first_m = (t - first_board_t) / 60.0

        # Rule 4: Full -> immediate takeoff
        if n_boarded >= EVTOL_CAPACITY:
            return self._dispatch(t)

        # Rule 2: First pax alone for 15 min
        if n_boarded == 1 and time_since_first_m >= FIRST_PAX_TIMEOUT_M:
            return self._dispatch(t)

        # Rule 3: 2+ pax, second pax waited > 10 min
        if n_boarded >= 2:
            second_board_t = self.boarded[1].board_time_s
            time_since_second_m = (t - second_board_t) / 60.0
            if time_since_second_m >= SECOND_PAX_TIMEOUT_M:
                return self._dispatch(t)

        return None

    def _dispatch(self, t: float) -> dict:
        """Take off with currently boarded passengers."""
        passengers = list(self.boarded)
        for pax in passengers:
            pax.takeoff_time_s = t
        self.departed.extend(passengers)
        self.boarded.clear()

        return {
            "origin_vp": self.origin_vp,
            "dest_vp": self.dest_vp,
            "passengers": len(passengers),
            "passenger_records": passengers,
            "takeoff_time_s": t,
        }


def load_passenger_queues() -> dict[str, VertiportQueue]:
    """
    Load passenger arrival data from the Excel file and create queues.
    Returns dict keyed by origin_vp.
    """
    if openpyxl is None:
        raise ImportError("openpyxl required — pip install openpyxl")

    wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True)

    # Tampa -> Brandon = vp_a -> vp_b
    vpa_times = []
    for row in wb["Tampa"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] == "Brandon":
            vpa_times.append(float(row[0]))

    # Brandon -> Tampa = vp_b -> vp_a
    vpb_times = []
    for row in wb["Brandon"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] == "Tampa":
            vpb_times.append(float(row[0]))

    wb.close()

    queues = {
        "vp_a": VertiportQueue("vp_a", "vp_b", vpa_times),
        "vp_b": VertiportQueue("vp_b", "vp_a", vpb_times),
    }

    print(f"[PassengerQueue] vp_a->vp_b: {len(vpa_times)} scheduled passengers")
    print(f"[PassengerQueue] vp_b->vp_a: {len(vpb_times)} scheduled passengers")
    return queues
